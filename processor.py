# -*- coding: utf-8 -*-
import os
import re
import json
import ast
from typing import Any, Optional, List, Callable

import pandas as pd
from openpyxl import Workbook
from openai import OpenAI


# ======== 可调参数（不改代码也能控制）========
DEFAULT_TIMEOUT_SEC = int(os.getenv("AI_TIMEOUT_SEC", "20"))     # 单次AI请求超时秒数
DEFAULT_RETRIES = int(os.getenv("AI_RETRIES", "2"))              # 单个sheet重试次数
MAX_SHEETS = int(os.getenv("MAX_SHEETS", "0"))                   # 0表示不限制；>0表示只处理前N个sheet（调试用）
PREVIEW_ROWS = int(os.getenv("AI_PREVIEW_ROWS", "15"))           # 传给模型的行数
PREVIEW_COLS = int(os.getenv("AI_PREVIEW_COLS", "25"))           # 传给模型的列数


def _parse_llm_list_of_lists(raw_text: str) -> List[List[Any]]:
    """尽量解析模型输出为二维列表(list[list])"""
    if raw_text is None:
        raise ValueError("模型返回为空")
    text = raw_text.strip()

    # 去掉 ``` 包裹
    if text.startswith("```"):
        text = text.strip("`").strip()

    # 先 JSON
    try:
        obj = json.loads(text)
        if isinstance(obj, list):
            return obj
    except Exception:
        pass

    # 再 Python literal
    try:
        obj = ast.literal_eval(text)
        if isinstance(obj, list):
            return obj
    except Exception:
        pass

    # 最后尝试“去转义再 loads”
    try:
        clean_text = text.replace('\\"', '"').replace('"\"', '"')
        obj = json.loads(clean_text)
        if isinstance(obj, list):
            return obj
    except Exception:
        pass

    raise ValueError(f"无法解析模型输出为二维列表。输出开头：{text[:200]}")


def _build_preview_csv(df: pd.DataFrame) -> str:
    """控制传给模型的内容大小，避免token爆炸"""
    if df is None or df.empty:
        return "（空表）"
    # 限制行列
    sub = df.iloc[:PREVIEW_ROWS, : min(PREVIEW_COLS, df.shape[1])]
    return sub.to_csv(index=False)


def analyze_data_with_llm(
    df: Any,
    api_key: Optional[str] = None,
    base_url: str = "https://dashscope.aliyuncs.com/compatible-mode/v1",
    model: str = "qwen3-max",
    timeout_sec: int = DEFAULT_TIMEOUT_SEC,
    retries: int = DEFAULT_RETRIES,
) -> List[List[Any]]:
    """调用通义千问（OpenAI 兼容接口）做映射清洗"""
    if api_key is None:
        api_key = os.getenv("DASHSCOPE_API_KEY")
    if not api_key:
        raise RuntimeError("未检测到环境变量 DASHSCOPE_API_KEY（请在WSGI或环境变量中配置）")

    client = OpenAI(api_key=api_key, base_url=base_url, timeout=timeout_sec)

    # 只传预览，避免超长
    if isinstance(df, pd.DataFrame):
        preview = _build_preview_csv(df)
    else:
        preview = str(df)[:5000]

    prompt = f"""
新建一个对话，你是一个专业的数据分析助手，请严格按照以下要求处理数据：
数据来源（CSV预览，含表头+部分数据）：\n{preview}

这个表前面行是表头，后面是数据，要从前面映射各产品，获取后面对应的数据。
一、数据清洗与预处理：
1. 数据筛选规则：
   - 不处理单位名称包含"未划分"、"合计"、"战客"的数据
   - 没有网格的：以"旗县"字段作为单位名称
   - 有网格的：以"网格"字段作为单位名称，但"旗县"仍需要输出为单独列
2. 指标映射规则（严格匹配顺序）：
| 目标字段 | 数据中的可能字段名（按匹配优先级降序） |
|----------|-------------------------------------|
| 日目标 | 日目标、日发展目标、当日目标、本日目标、日发展 |
| 日发展 | 日发展、当日发展、本日发展、日新增 |
| 日发展完成率 | 日完成率、日发展完成率、当日完成率；必须包含"日"且不包含"月" |
| 月目标 | 月目标、月度目标、本月目标 |
| 月累计发展 | 月累计发展、月度累计、累计发展、月发展 |
| 月完成率 | 月完成率、月度完成率、累计完成率；必须包含"月"或"累计"且不包含"日" |

映射规则：
- 按优先级匹配字段名
- 关键词必须满足要求（避免日/月混淆）
- 找不到对应字段输出 -1.5

二、输出格式要求：
- 只返回一个合法的二维列表(list of lists)，以 [[ 开始，以 ]] 结束，中间无额外内容
- 不要表头（只输出数据行）
- 每行严格为：["单位名称","日目标","日发展","日发展完成率","月目标","月累计发展","月完成率","得分","旗县"]
"""

    last_err: Optional[Exception] = None
    for attempt in range(1, retries + 1):
        try:
            completion = client.chat.completions.create(
                model=model,
                messages=[
                    {"role": "system", "content": "You are a helpful assistant."},
                    {"role": "user", "content": prompt},
                ],
            )
            content = completion.choices[0].message.content
            return _parse_llm_list_of_lists(content)
        except Exception as e:
            last_err = e
            # 最后一次失败就抛出
            if attempt == retries:
                raise
    # 理论到不了这里
    raise RuntimeError(f"AI调用失败：{last_err}")


def _read_excel_sheet(file_path: str, sheet_name: str, skiprows: int = 1, nrows: int = None, usecols: str = None):
    return pd.read_excel(
        file_path,
        sheet_name=sheet_name,
        skiprows=skiprows,
        nrows=nrows,
        usecols=usecols
    )


def process_excel(param_path: str, report_path: str, out_path: str, log_fn: Callable[[str], None] = lambda _: None) -> None:
    """
    网站入口：读取参数表 + 报表，调用 AI 做映射，写出 out_path
    注意：无论中途发生什么，都会在 finally 尝试写出 out_path，保证可下载
    """
    wb = Workbook()
    ws1 = wb.active

    # 表头（与队友输出对齐：包含产品/排名/备注列）
    all_headers = ["单位名称", "日目标", "日发展", "日发展完成率",
                   "月目标", "月累计发展", "月完成率", "得分", "旗县", "产品", "排名", "备注"]
    for col, header in enumerate(all_headers, start=1):
        ws1.cell(row=1, column=col, value=header)
    product_col_index = all_headers.index("产品") + 1
    hangshu = 2

    try:
        log_fn("读取参数表…")
        df_param_all = pd.read_excel(param_path)
        shifou = df_param_all.iloc[0, 1] if df_param_all.shape[1] > 1 else 0

        numbers_list = []
        letters_list = []
        canshu_list = []

        if shifou == 0:
            # A:C + nrows=4 + skiprows=1（按你队友逻辑）
            canshu = pd.read_excel(param_path, usecols="A:C", nrows=4, skiprows=1)
            canshu_list = canshu.values.tolist()

            for _, row in canshu.iterrows():
                row_numbers = []
                row_letters = []
                for col_index in [1, 2]:
                    item = row.iloc[col_index]
                    if pd.isna(item):
                        numbers = ''
                        letters = ''
                    else:
                        item_str = str(item)
                        numbers = ''.join(re.findall(r'\d+', item_str))
                        letters = ''.join(re.findall(r'[a-zA-Z]+', item_str))
                    row_numbers.append(numbers)
                    row_letters.append(letters)
                numbers_list.append(row_numbers)
                letters_list.append(row_letters)
        else:
            xls_report = pd.ExcelFile(report_path)
            canshu_list = [[s] for s in xls_report.sheet_names]

        xls = pd.ExcelFile(report_path)
        sheet_names = xls.sheet_names
        log_fn(f"报表包含 {len(sheet_names)} 个sheet")

        # 遍历参数表指定的 sheet
        for i in range(len(canshu_list)):
            if MAX_SHEETS > 0 and i >= MAX_SHEETS:
                log_fn(f"MAX_SHEETS={MAX_SHEETS}，停止继续处理（调试模式）")
                break

            target_sheet = canshu_list[i][0]
            if target_sheet not in sheet_names:
                log_fn(f"跳过：sheet不存在 {target_sheet}")
                continue

            log_fn(f"处理sheet：{target_sheet}")

            # 读表
            if shifou == 0:
                try:
                    number = int(numbers_list[i][1]) - int(numbers_list[i][0]) - 4
                    letter1 = letters_list[i][0]
                    letter2 = letters_list[i][1]
                    range_str = f"{letter1}:{letter2}"
                except Exception as e:
                    log_fn(f"参数表范围解析失败：{type(e).__name__}: {e}，跳过该sheet")
                    continue

                df = _read_excel_sheet(
                    file_path=report_path,
                    sheet_name=target_sheet,
                    skiprows=4,
                    nrows=number,
                    usecols=range_str,
                )
            else:
                df = _read_excel_sheet(
                    file_path=report_path,
                    sheet_name=target_sheet,
                    skiprows=1
                )

            # 调 AI（关键：有超时 + 有日志 + 失败不中断整体）
            log_fn("准备调用AI…")
            try:
                all_data = analyze_data_with_llm(df)
                log_fn(f"AI返回成功 ✅ 行数={len(all_data)}")
            except Exception as e:
                log_fn(f"AI调用失败：{type(e).__name__}: {e}")
                all_data = []

            # 写入
            for row_data in all_data:
                for col_idx, value in enumerate(row_data, 1):
                    ws1.cell(row=hangshu, column=col_idx, value=value)
                ws1.cell(row=hangshu, column=product_col_index, value=target_sheet)
                hangshu += 1

    finally:
        # 无论怎样都尝试写出，保证下载不再提示“尚未生成”
        log_fn("写出结果文件…")
        # outputs 目录不存在也创建一下（双保险）
        os.makedirs(os.path.dirname(out_path), exist_ok=True)
        wb.save(out_path)
        log_fn("写出完成 ✅")